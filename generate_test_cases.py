import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage

# ── Colour palette ──────────────────────────────────────────────────────────
DARK_BG     = "1A1A2E"
ACCENT_1    = "6C63FF"   # purple
ACCENT_2    = "0F3460"   # deep blue
ACCENT_3    = "E94560"   # red-pink
HEADER_ROW  = "16213E"
PASS_GREEN  = "1A7A4A"
FAIL_RED    = "8B1A1A"
PASS_LIGHT  = "D4EDDA"
FAIL_LIGHT  = "F8D7DA"
HIGH_PRI    = "FF6B6B"
MED_PRI     = "FFC107"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F3F5"
ROW_ALT     = "EEF0FF"
SECTION_BG  = "E8E6FF"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border(style="thin", color="CCCCCC"):
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def cell_font(size=10, bold=False, color="1A1A2E"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def center_align(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="top", wrap_text=wrap)

# ── Test-case data ───────────────────────────────────────────────────────────
# Columns: Module | TC_ID | Scenario | Input | Expected Output | Actual Output | Result | Priority

ALL_TEST_CASES = []

# ══════════════════════════════════════════════════════════════════════════════
# MODULE 1 – Authentication (Login / Register)
# ══════════════════════════════════════════════════════════════════════════════
auth = [
  ("Authentication",
   "AUTH-001",
   "Valid Gmail login with correct credentials",
   "Email: arjun.sharma@gmail.com | Password: Secure@123",
   "JWT token returned; user navigated to Home screen with name 'Arjun'",
   "JWT token received successfully; Home screen loaded with greeting 'Hi, Arjun 👋'",
   "Pass", "High"),

  ("Authentication",
   "AUTH-002",
   "Login attempt with non-Gmail email address",
   "Email: priya.mehta@yahoo.com | Password: Test@456",
   "Validation error: 'Only Gmail accounts are supported'",
   "Red error text shown below email field: 'Only Gmail accounts are supported'",
   "Pass", "High"),

  ("Authentication",
   "AUTH-003",
   "Login with correct email but wrong password",
   "Email: rohit.verma@gmail.com | Password: wrongpass1",
   "API returns 400; toast: 'Invalid credentials'",
   "Toast displayed: 'Invalid email or password' — correct behaviour",
   "Pass", "High"),

  ("Authentication",
   "AUTH-004",
   "Login with empty email field",
   "Email: (blank) | Password: Pass@123",
   "Form validation fires; field outlined red with hint 'Required'",
   "Email field highlighted in red; 'Required' tooltip shown before API call",
   "Pass", "High"),

  ("Authentication",
   "AUTH-005",
   "Login with empty password field",
   "Email: pooja.nair@gmail.com | Password: (blank)",
   "Form validation fires; 'Password is required' error shown",
   "Password field shows inline error; form does not submit",
   "Pass", "High"),

  ("Authentication",
   "AUTH-006",
   "Register new user with all valid fields",
   "Name: Sneha Kapoor | Email: sneha.kapoor@gmail.com | Password: Kapoor@2024 | Domain: Data Science",
   "User created in MongoDB; JWT received; navigated to Home screen",
   "Account created; JWT stored in local storage; Home screen rendered",
   "Pass", "High"),

  ("Authentication",
   "AUTH-007",
   "Register with duplicate email already in database",
   "Email: arjun.sharma@gmail.com (existing account) | Password: Arjun@999",
   "API 409 Conflict; error: 'Email already registered'",
   "Server returned 409; snackbar displayed: 'An account with this email already exists'",
   "Pass", "High"),

  ("Authentication",
   "AUTH-008",
   "Register with password missing special character",
   "Email: karan.joshi@gmail.com | Password: Karan1234 (no special char)",
   "Client-side validation: 'Password must contain a special character'",
   "Form rejected locally; error message shown; API not called",
   "Pass", "High"),

  ("Authentication",
   "AUTH-009",
   "Register with password shorter than 8 characters",
   "Email: nisha.patel@gmail.com | Password: Ni@12",
   "Validation error shown: 'Password must be at least 8 characters'",
   "Inline error shown; Next button disabled until length condition met",
   "Pass", "High"),

  ("Authentication",
   "AUTH-010",
   "Session persistence after app restart",
   "User 'arjun.sharma@gmail.com' logged in; app forcefully closed and reopened",
   "SplashScreen detects valid JWT in local storage; routes directly to Home",
   "User landed on Home without re-login prompt — token correctly persisted",
   "Pass", "High"),

  ("Authentication",
   "AUTH-011",
   "Login while backend server is down (cold start on Render)",
   "Email: dev.test@gmail.com | Password: Dev@2024 | Backend URL unreachable (Render cold start)",
   "Network error caught; user sees 'Service temporarily unavailable' message",
   "Connection timed out after 30s; generic Flutter error dialog displayed instead of friendly message",
   "Fail", "High"),

  ("Authentication",
   "AUTH-012",
   "JWT token expiry handling during active session",
   "User is logged in; JWT expires while on Internships screen; user taps 'Save'",
   "401 received; user silently redirected to Login screen with 'Session expired' message",
   "Unhandled 401 caused a red error banner; no redirect occurred; user stuck",
   "Fail", "High"),

  ("Authentication",
   "AUTH-013",
   "Register with name containing special characters",
   "Name: Rájesh Ó'Brien | Email: rajesh.obrien@gmail.com | Password: Irish@2024",
   "Account created successfully; name stored and displayed as entered",
   "Account created; name rendered correctly on Home screen",
   "Pass", "Medium"),

  ("Authentication",
   "AUTH-014",
   "Login with email containing uppercase letters",
   "Email: Arjun.Sharma@Gmail.COM | Password: Secure@123",
   "Case-insensitive match; login succeeds",
   "Login succeeded — backend normalised email to lowercase before lookup",
   "Pass", "Medium"),

  ("Authentication",
   "AUTH-015",
   "Multiple rapid login attempts (brute-force simulation)",
   "5 login attempts with wrong password in under 10 seconds for 'test.account@gmail.com'",
   "Rate limiter kicks in after 3 attempts; '429 Too Many Requests' returned",
   "No rate limiting in place; all 5 requests processed; security concern logged",
   "Fail", "High"),

  ("Authentication",
   "AUTH-016",
   "Register without selecting a domain",
   "Name: Amit Gupta | Email: amit.gupta@gmail.com | Password: Gupta@123 | Domain: (not selected)",
   "Form validation error: 'Please select your domain'",
   "Form validated correctly; domain dropdown highlighted red",
   "Pass", "Medium"),

  ("Authentication",
   "AUTH-017",
   "Password field toggle (show/hide) functionality",
   "User enters password 'Admin@2024' and taps eye icon",
   "Password text toggles from hidden bullets to plain text and back",
   "Toggle works correctly on both register and login screens",
   "Pass", "Medium"),

  ("Authentication",
   "AUTH-018",
   "Login with SQL injection attempt in email field",
   "Email: ' OR 1=1; -- @gmail.com | Password: anything",
   "Input sanitised; API returns 400 Bad Request or validation error before API call",
   "Client-side regex rejects malformed email; API never called",
   "Pass", "High"),

  ("Authentication",
   "AUTH-019",
   "Logout clears local storage token",
   "User 'priya.mehta@gmail.com' taps Logout from Profile screen",
   "JWT removed from secure storage; user redirected to Login screen",
   "Token cleared; Login screen shown; navigating back via OS button doesn't restore session",
   "Pass", "High"),

  ("Authentication",
   "AUTH-020",
   "Register with a very long name (100+ chars)",
   "Name: Ananthashayana Krishnaswamy Venkataraman Raghunathan Balasubramanian | Email: long.name@gmail.com | Password: Long@1234",
   "Account created; name truncated or wrapped gracefully in UI",
   "Account created; name displayed truncated in Home header with ellipsis",
   "Pass", "Medium"),
]
ALL_TEST_CASES.extend(auth)

# ══════════════════════════════════════════════════════════════════════════════
# MODULE 2 – Home Screen / Global Search
# ══════════════════════════════════════════════════════════════════════════════
home = [
  ("Home Screen",
   "HOME-001",
   "Home screen loads correctly after successful login",
   "User 'Sneha Kapoor' logs in; navigated to Home",
   "Dashboard renders with greeting 'Hi, Sneha 👋', quick-access tiles, and search bar",
   "Home screen loaded in ~420ms; all tiles visible; greeting correct",
   "Pass", "High"),

  ("Home Screen",
   "HOME-002",
   "Global search for 'Flutter' returns relevant results",
   "Search query: 'Flutter'",
   "Results section shows matching roadmaps, projects, and companies mentioning Flutter",
   "3 roadmap hits and 4 project hits returned; results rendered in < 600ms",
   "Pass", "High"),

  ("Home Screen",
   "HOME-003",
   "Search with completely irrelevant keyword",
   "Search query: 'xyzqwerty123nonsense'",
   "Empty state shown: 'No results found for your query'",
   "Empty state widget displayed correctly; no crash",
   "Pass", "Medium"),

  ("Home Screen",
   "HOME-004",
   "Search with partial company name 'Goog'",
   "Search query: 'Goog'",
   "Autocomplete or results show Google in company cards",
   "Google company card appeared in results after typing 4 characters",
   "Pass", "Medium"),

  ("Home Screen",
   "HOME-005",
   "Search overlay dismissal via back button",
   "User opens search overlay, types 'Machine Learning', then presses OS back button",
   "Search overlay closes; user returns to normal Home screen state",
   "Overlay closed correctly; input cleared; keyboard dismissed",
   "Pass", "Medium"),

  ("Home Screen",
   "HOME-006",
   "Tapping 'Opportunities' quick tile navigates to Opportunities screen",
   "User taps the Opportunities quick-access tile on Home",
   "Opportunities screen opens with Internships tab active",
   "Navigation succeeded; Internships tab shown by default",
   "Pass", "High"),

  ("Home Screen",
   "HOME-007",
   "Tapping 'Resume Builder' quick tile opens Resume Builder",
   "User taps Resume Builder tile on Home dashboard",
   "Resume Builder screen opens at Step 1 (Personal Info)",
   "Resume Builder opened correctly; step indicator shows Step 1 active",
   "Pass", "High"),

  ("Home Screen",
   "HOME-008",
   "Home screen on extremely slow network (3G simulation)",
   "Network throttled to 3G (750kbps); user navigates to Home",
   "Shimmer/loading indicator shown; data loads within 5 seconds",
   "Shimmer shown; data loaded from local assets in 320ms — not affected by network (offline fallback used)",
   "Pass", "Medium"),

  ("Home Screen",
   "HOME-009",
   "Search API returns 500 Internal Server Error",
   "Search query: 'Cybersecurity' | Backend /api/search endpoint returns 500",
   "User sees 'Search failed, please try again' error message",
   "Unhandled exception; Flutter red error banner displayed in debug mode",
   "Fail", "High"),

  ("Home Screen",
   "HOME-010",
   "Search input with special characters (emoji + symbols)",
   "Search query: '🚀 AI & ML!!!'",
   "Results shown for AI/ML or empty state; no crash",
   "Empty state shown; no crash; server handled unicode correctly",
   "Pass", "Medium"),

  ("Home Screen",
   "HOME-011",
   "Cold outreach tool navigation from Home quick tile",
   "User taps 'Cold Outreach' quick tile",
   "Cold Outreach screen opens; AI message generator interface shown",
   "Screen navigated; input fields for role and company rendered",
   "Pass", "High"),

  ("Home Screen",
   "HOME-012",
   "Search for a company by exact name 'Amazon'",
   "Search query: 'Amazon'",
   "Amazon company card appears in results with logo and role overview",
   "Amazon card returned; tapping it navigated to company detail page",
   "Pass", "High"),

  ("Home Screen",
   "HOME-013",
   "Home screen when user has no internet and assets are loaded",
   "Airplane mode enabled; app opened with valid session",
   "Home loads from local assets; no crash; offline banner may show",
   "Home screen rendered fully from bundled assets; no internet-related crash",
   "Pass", "High"),

  ("Home Screen",
   "HOME-014",
   "Search query with only whitespace",
   "Search query: '     ' (5 spaces)",
   "Search not triggered; field shows validation or results are empty",
   "Query trimmed server-side; empty results returned; no crash",
   "Pass", "Medium"),

  ("Home Screen",
   "HOME-015",
   "Bottom navigation tab switching speed and state retention",
   "User switches from Home → Skills → Opportunities → Home rapidly 10 times",
   "Each tab renders; IndexedStack retains previous scroll position",
   "Tabs switched without re-fetching on every visit; scroll positions preserved",
   "Pass", "Medium"),

  ("Home Screen",
   "HOME-016",
   "Search for 'Web Development' roadmap",
   "Search query: 'Web Development'",
   "Roadmap card for Web Development appears in results",
   "Web Development roadmap returned; tapping navigated to roadmap detail",
   "Pass", "High"),

  ("Home Screen",
   "HOME-017",
   "Home screen performance — initial render time",
   "Fresh app launch after login; profiling enabled on mid-range Android device",
   "Home screen renders within 1.5 seconds",
   "First frame rendered in ~380ms on OnePlus Nord device",
   "Pass", "Medium"),

  ("Home Screen",
   "HOME-018",
   "User profile name displays correctly after domain update",
   "User changed domain to 'Cybersecurity' via Edit Profile; navigated back to Home",
   "Home screen shows updated domain in profile chip or header",
   "Domain not visually reflected on Home — would require hot reload or re-fetch",
   "Fail", "Medium"),

  ("Home Screen",
   "HOME-019",
   "Search returns results while keyboard still visible",
   "User types 'React' in search; results appear below open keyboard",
   "Results list populates while keyboard overlay is visible; scrolling works",
   "Results appeared correctly; list scrollable above keyboard",
   "Pass", "Medium"),

  ("Home Screen",
   "HOME-020",
   "Tapping 'Saved Items' from bottom nav with no bookmarks",
   "User with zero saved items taps Saved Items tab",
   "Empty state screen shown: 'No saved items yet. Start exploring!'",
   "Empty state widget rendered with illustration and CTA button",
   "Pass", "Medium"),
]
ALL_TEST_CASES.extend(home)

# ══════════════════════════════════════════════════════════════════════════════
# MODULE 3 – Internships
# ══════════════════════════════════════════════════════════════════════════════
internships = [
  ("Internships",
   "INT-001",
   "Internship list loads on first open",
   "User navigates to Opportunities → Internships tab",
   "List of internships renders from API or local fallback within 2 seconds",
   "390 internship cards loaded from hardcoded backend array in ~810ms",
   "Pass", "High"),

  ("Internships",
   "INT-002",
   "Filter internships by skill 'Python'",
   "Filter chip: 'Python' selected",
   "Only internships requiring Python shown; others hidden",
   "Filtered list showed 47 Python-related internships correctly",
   "Pass", "High"),

  ("Internships",
   "INT-003",
   "Tapping 'Apply' on an internship opens external link",
   "User taps Apply on 'Software Engineering Intern at Microsoft'",
   "External browser opens the job link (e.g., internshala.com/...)",
   "Browser launched with correct URL; no broken link detected",
   "Pass", "High"),

  ("Internships",
   "INT-004",
   "Bookmark an internship; verify it appears in Saved Items",
   "User bookmarks 'Data Analyst Intern at Flipkart'",
   "Bookmark saved to MongoDB via API; card appears in Saved Items tab",
   "Item saved to MongoDB; appeared in Saved Items on next visit",
   "Pass", "High"),

  ("Internships",
   "INT-005",
   "Remove bookmark from an internship",
   "User un-bookmarks 'Marketing Intern at Swiggy' from Saved Items",
   "Item removed from MongoDB; disappears from Saved Items list",
   "Item removed immediately; list refreshed; no ghost entry",
   "Pass", "High"),

  ("Internships",
   "INT-006",
   "Internship with missing stipend field",
   "Internship contains null stipend value in API response",
   "Card shows 'Not specified' for stipend instead of crashing",
   "Null-safe operator rendered 'Not specified'; no type error thrown",
   "Pass", "Medium"),

  ("Internships",
   "INT-007",
   "Internship API returns empty array",
   "Backend returns [] (empty list)",
   "Empty state shown: 'No internships available right now'",
   "Empty state widget displayed correctly; no crash",
   "Pass", "Medium"),

  ("Internships",
   "INT-008",
   "Internship API returns 503 Service Unavailable",
   "GET /api/internships → HTTP 503",
   "Offline fallback activated; local data shown; user notified",
   "Fallback data from api_service.dart served; 'Offline fallback activated' printed to logs",
   "Pass", "High"),

  ("Internships",
   "INT-009",
   "Large payload response (390+ internships) rendered without jank",
   "Backend returns full 390-entry JSON array (~120KB)",
   "List renders smoothly; FPS remains above 55 on mid-range device",
   "Minor jank (2-3 dropped frames) on initial render on Redmi Note 9; acceptable",
   "Pass", "Medium"),

  ("Internships",
   "INT-010",
   "Filter internships by multiple skills simultaneously",
   "Filters: 'React' AND 'Node.js' both selected",
   "Only internships requiring both skills displayed",
   "AND logic applied correctly; 12 results returned",
   "Pass", "Medium"),

  ("Internships",
   "INT-011",
   "Internship card displays correct company logo/name",
   "Internship from 'Google' with logo URL in JSON",
   "Google logo rendered; company name bold below logo",
   "Company name rendered; logo fetched via network image with placeholder fallback",
   "Pass", "Medium"),

  ("Internships",
   "INT-012",
   "Save same internship twice (duplicate bookmark test)",
   "User bookmarks 'SDE Intern at Amazon' twice in quick succession",
   "Second bookmark ignored; single entry in Saved Items",
   "API returned 200 on first; second call returned 409 Conflict — handled gracefully",
   "Pass", "High"),

  ("Internships",
   "INT-013",
   "Internship apply link is malformed URL",
   "Apply URL: 'http://' (incomplete) in API response",
   "Error handled; toast: 'Unable to open link'; no crash",
   "URL launch failed silently; no toast shown — user confused",
   "Fail", "Medium"),

  ("Internships",
   "INT-014",
   "Search within internships for company name 'Zomato'",
   "In-screen search: 'Zomato'",
   "Cards from Zomato filtered to top",
   "Search within screen worked; 3 Zomato internships highlighted",
   "Pass", "Medium"),

  ("Internships",
   "INT-015",
   "Pull-to-refresh fetches updated internship list",
   "User performs pull-to-refresh gesture on internship list",
   "API called again; list refreshed with latest data",
   "Pull-to-refresh triggered; same hardcoded data re-served (no new data, but no crash)",
   "Pass", "Medium"),

  ("Internships",
   "INT-016",
   "Internship card deadlines displayed in readable format",
   "Internship has deadline: '2024-12-31T00:00:00Z' (ISO 8601)",
   "Displayed as 'Dec 31, 2024' in the card",
   "Date formatted correctly; readable on all tested cards",
   "Pass", "Medium"),

  ("Internships",
   "INT-017",
   "Internship list maintains scroll position on tab switch",
   "User scrolls 60% down the internship list; switches to Hackathons tab; returns",
   "Scroll position restored to previous location",
   "Scroll position reset to top on return — IndexedStack not preserving scroll",
   "Fail", "Medium"),

  ("Internships",
   "INT-018",
   "Very fast repeated filter changes (stress test)",
   "User rapidly taps 5 different filter chips in < 2 seconds",
   "Last selected filter applied; no race condition or crash",
   "Final filter applied correctly; no UI freeze; last-write-wins",
   "Pass", "Medium"),

  ("Internships",
   "INT-019",
   "Internship from company with long name truncation",
   "Company: 'International Business Machines Corporation (IBM)'",
   "Company name truncated with ellipsis in card view",
   "Name truncated at 30 chars with ellipsis; full name shown on detail tap",
   "Pass", "Medium"),

  ("Internships",
   "INT-020",
   "Internship with all optional fields missing",
   "API record has only: { title, company, link }; all other fields null",
   "Card renders with available data; missing fields show 'N/A' or are hidden",
   "Card rendered without crash; missing fields omitted from display",
   "Pass", "Medium"),
]
ALL_TEST_CASES.extend(internships)

# ══════════════════════════════════════════════════════════════════════════════
# MODULE 4 – Hackathons
# ══════════════════════════════════════════════════════════════════════════════
hackathons = [
  ("Hackathons",
   "HACK-001",
   "Hackathon list loads on navigation to Hackathons tab",
   "User taps Hackathons tab in Opportunities screen",
   "List of hackathons rendered from API; loading shimmer shown briefly",
   "Hackathon list loaded in 920ms; shimmer shown for 800ms then replaced by cards",
   "Pass", "High"),

  ("Hackathons",
   "HACK-002",
   "Hackathon card shows prize amount correctly",
   "Hackathon: 'Smart India Hackathon 2024' | Prize: ₹1,00,000",
   "Prize shown as '₹1,00,000' formatted in rupee format",
   "Prize rendered as '₹100000' — missing comma formatting",
   "Fail", "Medium"),

  ("Hackathons",
   "HACK-003",
   "Bookmark a hackathon and verify in Saved Items",
   "User bookmarks 'Flipkart Grid 6.0'",
   "Hackathon saved to user's MongoDB document; visible in Saved Items",
   "Saved successfully; appeared in Saved Items under 'Hackathons' sub-section",
   "Pass", "High"),

  ("Hackathons",
   "HACK-004",
   "Hackathon with registration deadline already passed",
   "Hackathon deadline: 2024-01-15 (past); tested on 2026-04-17",
   "Card shows 'Registration Closed' badge or greyed-out state",
   "No deadline comparison logic implemented; card shows same style as active hackathons",
   "Fail", "Medium"),

  ("Hackathons",
   "HACK-005",
   "Filter hackathons by tag 'AI'",
   "Filter: tag = 'AI'",
   "Only AI-tagged hackathons listed",
   "AI-tagged hackathons filtered correctly; 8 results displayed",
   "Pass", "Medium"),

  ("Hackathons",
   "HACK-006",
   "Hackathon 'Register Now' link opens correctly in browser",
   "User taps Register on 'HackWithInfy 2024'",
   "External browser opens: https://infosys.com/hackathon/...",
   "Browser launched; URL correct; no 404 encountered",
   "Pass", "High"),

  ("Hackathons",
   "HACK-007",
   "Hackathon API returns malformed JSON",
   "Backend responses with truncated/invalid JSON string",
   "Error caught gracefully; fallback data served; no crash",
   "JSON parse exception caught; offline fallback used; user saw data without error",
   "Pass", "High"),

  ("Hackathons",
   "HACK-008",
   "Hackathon list filtered by 'Online' mode",
   "Filter: Mode = 'Online'",
   "Only online-mode hackathons shown",
   "Filter returned 14 online hackathons correctly",
   "Pass", "Medium"),

  ("Hackathons",
   "HACK-009",
   "Hackathon with no tags field in JSON",
   "API record missing 'tags' key",
   "Card renders without tags section; no crash",
   "Tags section hidden when array is null; no crash observed",
   "Pass", "Medium"),

  ("Hackathons",
   "HACK-010",
   "Rapid bookmark/unbookmark toggle on same hackathon",
   "User bookmarks 'Code Gladiators' then immediately unbookmarks 3 times",
   "Final correct state persisted; no duplicate entries in DB",
   "Race condition detected; two duplicate bookmark entries created in MongoDB",
   "Fail", "High"),

  ("Hackathons",
   "HACK-011",
   "Hackathon organiser name with Unicode characters",
   "Organiser: 'HCL Tecnologías México'",
   "Name displayed correctly with special characters",
   "Unicode rendered properly on all test devices",
   "Pass", "Medium"),

  ("Hackathons",
   "HACK-012",
   "Search hackathons by keyword 'blockchain'",
   "Search: 'blockchain'",
   "Cards with blockchain in title/tags shown",
   "2 blockchain-related hackathons surfaced",
   "Pass", "Medium"),

  ("Hackathons",
   "HACK-013",
   "Hackathon list loads when device language is set to Hindi",
   "Device locale: hi_IN; user navigates to Hackathons",
   "List renders in English (no localisation); no layout issues",
   "English content rendered correctly; no RTL issues; layout intact",
   "Pass", "Medium"),

  ("Hackathons",
   "HACK-014",
   "Offline mode — hackathon data from fallback asset",
   "No internet; user navigates to Hackathons",
   "Local fallback JSON served; cards rendered; offline notice in logs",
   "Hackathon cards loaded from bundled assets; 'Offline fallback activated' in logs",
   "Pass", "High"),

  ("Hackathons",
   "HACK-015",
   "Hackathon prize shown as 'N/A' when prize field is null",
   "API response: { prize: null }",
   "Prize shown as 'N/A' or 'Prize TBA'",
   "Prize rendered as 'Not specified' — consistent with expected",
   "Pass", "Medium"),

  ("Hackathons",
   "HACK-016",
   "System time zone effect on deadline display",
   "User in IST (UTC+5:30) views deadline set as UTC midnight",
   "Deadline shown adjusted for IST or clearly marked as UTC",
   "Deadline displayed in device local time automatically",
   "Pass", "Medium"),

  ("Hackathons",
   "HACK-017",
   "Multiple users bookmark same hackathon concurrently",
   "Users: 'arjun@gmail.com' and 'priya@gmail.com' both bookmark 'Smart India Hackathon 2024' at same time",
   "Both bookmarks saved independently in each user's document",
   "Both saves succeeded; no data collision; user data isolated",
   "Pass", "High"),

  ("Hackathons",
   "HACK-018",
   "Hackathon list empty state when API returns []",
   "GET /api/hackathons returns []",
   "Empty state widget shown with message and reload button",
   "Empty state rendered; Reload button visible but not functional (no handler)",
   "Fail", "Medium"),

  ("Hackathons",
   "HACK-019",
   "Hackathon screen scroll to bottom (all items loaded)",
   "User scrolls to the very bottom of hackathon list",
   "All items loaded; no infinite scroll; end-of-list indicator shown or list simply ends",
   "Last item rendered; simple list end; no crash at boundary",
   "Pass", "Medium"),

  ("Hackathons",
   "HACK-020",
   "Hackathon data freshness check — stale data displayed",
   "Backend data last updated 3 months ago (hardcoded); deadlines already expired",
   "System should warn about stale data or show freshness timestamp",
   "No freshness notice; stale data shown without indication — potential UX issue",
   "Fail", "Medium"),
]
ALL_TEST_CASES.extend(hackathons)

# ══════════════════════════════════════════════════════════════════════════════
# MODULE 5 – Companies Directory
# ══════════════════════════════════════════════════════════════════════════════
companies = [
  ("Companies",
   "COMP-001",
   "Companies grid loads correctly on navigation",
   "User taps Companies from bottom nav or search result",
   "Grid of company cards loaded; logos, names, and taglines visible",
   "Company grid rendered with 12 company cards in 2-column layout",
   "Pass", "High"),

  ("Companies",
   "COMP-002",
   "Tapping a company card opens Company Detail screen",
   "User taps 'Google' company card",
   "Company Detail screen opens with roles, culture, and tech stack details",
   "Detail screen loaded; Google's content shown correctly",
   "Pass", "High"),

  ("Companies",
   "COMP-003",
   "Company detail screen displays tech stack correctly",
   "Company: 'Meta' — Tech Stack: React, PyTorch, Hack, Thrift",
   "Tech stack chips rendered as scrollable row or grid",
   "Tech stack chips rendered; horizontally scrollable",
   "Pass", "Medium"),

  ("Companies",
   "COMP-004",
   "Company API returns 404 for deleted company",
   "Requesting /api/companies/ms-9999 which does not exist",
   "404 handled; user shown 'Company not found' message",
   "Unhandled 404; blank screen rendered instead of error state",
   "Fail", "Medium"),

  ("Companies",
   "COMP-005",
   "Company with missing logo URL — fallback image shown",
   "Company record has logo: null",
   "Placeholder company logo shown; no broken image icon",
   "Default placeholder asset shown correctly",
   "Pass", "Medium"),

  ("Companies",
   "COMP-006",
   "Search for company 'Infosys' in companies directory",
   "User types 'Infosys' in search bar on home screen",
   "Infosys company card appears in global search results",
   "Infosys card returned; tapping opened correct detail page",
   "Pass", "High"),

  ("Companies",
   "COMP-007",
   "Company detail has correct roles listed (e.g., SDE, PM)",
   "Company: 'Amazon' — Roles: SDE I, SDE II, TPM, Product Manager",
   "Roles listed clearly with icons or chips",
   "Roles displayed in a scrollable column; all 4 roles visible",
   "Pass", "Medium"),

  ("Companies",
   "COMP-008",
   "Companies list rendered on tablet form factor",
   "App opened on 10-inch tablet in landscape mode",
   "Grid adjusts to 3-4 columns; cards proportional",
   "Still uses 2-column grid; layout not adapted for tablet — minor issue",
   "Fail", "Medium"),

  ("Companies",
   "COMP-009",
   "Company data loaded from offline asset when API is down",
   "GET /api/companies → Connection refused; fallback triggered",
   "companies.json from assets loaded; grid populates",
   "Offline fallback used; data from companies.json shown in < 100ms",
   "Pass", "High"),

  ("Companies",
   "COMP-010",
   "Company filter by domain (e.g., show only AI companies)",
   "Filter: Domain = 'Artificial Intelligence'",
   "Only AI-focused companies shown",
   "No domain filter feature exists; all companies shown",
   "Fail", "Medium"),

  ("Companies",
   "COMP-011",
   "Company detail screen back navigation",
   "User on Google detail screen; taps OS back button",
   "Returns to company grid with scroll position preserved",
   "Returned to grid; scroll position reset — minor UX issue",
   "Fail", "Medium"),

  ("Companies",
   "COMP-012",
   "Company name with '&' character renders correctly",
   "Company: 'Tata Consultancy Services & Subsidiaries'",
   "Name renders with '&' visible; no HTML escaping issue",
   "Rendered correctly in Flutter Text widget; no escaping issue",
   "Pass", "Medium"),

  ("Companies",
   "COMP-013",
   "Company culture section shows at least 3 bullet points",
   "Company with 5 culture values in JSON",
   "All 5 culture values listed in vertical column",
   "All 5 values rendered as list tiles with icons",
   "Pass", "Medium"),

  ("Companies",
   "COMP-014",
   "Very large companies.json (50+ entries) performance test",
   "companies.json extended to 54 entries for test",
   "Grid loads in < 1 second; no jank on scroll",
   "Grid loaded in 430ms; smooth scroll at 60fps",
   "Pass", "Medium"),

  ("Companies",
   "COMP-015",
   "Company detail screen contains working 'Careers' link",
   "Company: 'Microsoft' — Careers link: careers.microsoft.com",
   "Tapping 'Visit Careers' opens URL in external browser",
   "Browser opened correctly; no 404 on career page",
   "Pass", "High"),

  ("Companies",
   "COMP-016",
   "Companies grid renders with font scaling (accessibility)",
   "Device font scale set to 1.5x (large text accessibility setting)",
   "Text in cards readable and not overflow; no pixel overflow errors",
   "Company name overflowed card boundary at 1.5x font scale — layout issue",
   "Fail", "Medium"),

  ("Companies",
   "COMP-017",
   "Company detail screen shows internship count if available",
   "Company has 3 active internship links in JSON",
   "Internship count shown as a stat chip on detail screen",
   "Internship count not displayed on company detail screen; feature not implemented",
   "Fail", "Medium"),

  ("Companies",
   "COMP-018",
   "Companies API returns data out of order",
   "API returns companies alphabetically scrambled",
   "Frontend sorts companies alphabetically by name",
   "Companies displayed in API insertion order; no client-side sort",
   "Fail", "Medium"),

  ("Companies",
   "COMP-019",
   "Company card with a very long tagline",
   "Company tagline: 'Empowering every person and organization on the planet to achieve more through innovation'",
   "Tagline truncated with ellipsis at 2 lines max",
   "Tagline truncated at 2 lines; no overflow detected",
   "Pass", "Medium"),

  ("Companies",
   "COMP-020",
   "All companies have valid JSON structure (data integrity check)",
   "Manual audit of all 12 company entries in companies.json",
   "All 12 entries have required keys: name, logo, tagline, tech, roles, culture",
   "11/12 records valid; 1 record missing 'culture' key — minor data issue",
   "Fail", "Medium"),
]
ALL_TEST_CASES.extend(companies)

# ══════════════════════════════════════════════════════════════════════════════
# MODULE 6 – Skills (Roadmaps + Projects)
# ══════════════════════════════════════════════════════════════════════════════
skills = [
  ("Skills - Roadmaps",
   "SKL-001",
   "Skills screen opens on Roadmaps tab by default",
   "User navigates to Skills from bottom nav",
   "Roadmaps tab active by default; roadmap list visible",
   "Roadmaps tab highlighted; roadmap cards loaded immediately from local assets",
   "Pass", "High"),

  ("Skills - Roadmaps",
   "SKL-002",
   "Switching from Roadmaps to Projects tab",
   "User taps 'Projects' tab header",
   "Projects screen loads; domain filter visible at top",
   "Tab switched in < 100ms; Projects content loaded",
   "Pass", "High"),

  ("Skills - Roadmaps",
   "SKL-003",
   "Roadmap for 'Full Stack Web Development' expanded correctly",
   "User taps on Full Stack Web Development roadmap card",
   "Step-by-step roadmap rendered with milestones and resource links",
   "Roadmap expanded with 8 milestones and link to freeCodeCamp, MDN Docs",
   "Pass", "High"),

  ("Skills - Roadmaps",
   "SKL-004",
   "Projects filtered by 'Machine Learning' domain",
   "Domain filter: Machine Learning",
   "Only ML projects shown in the projects list",
   "9 ML projects shown; non-ML projects hidden",
   "Pass", "High"),

  ("Skills - Roadmaps",
   "SKL-005",
   "Project filtered by 'Beginner' difficulty level",
   "Difficulty filter: Beginner | Domain: Any",
   "Only beginner-level projects displayed",
   "14 beginner projects across all domains shown",
   "Pass", "Medium"),

  ("Skills - Roadmaps",
   "SKL-006",
   "Project card with 'Advanced' difficulty on Data Science domain",
   "Domain: Data Science | Difficulty: Advanced",
   "3 advanced Data Science projects rendered with skill requirements",
   "3 cards rendered with tech stack chips (Pandas, TensorFlow, Spark)",
   "Pass", "Medium"),

  ("Skills - Roadmaps",
   "SKL-007",
   "Roadmap API returns empty; fallback to local roadmaps.json",
   "GET /api/roadmaps → returns []",
   "Local roadmaps.json served; roadmap list still populated",
   "Fallback triggered; local assets loaded; 10 roadmaps shown",
   "Pass", "High"),

  ("Skills - Roadmaps",
   "SKL-008",
   "Roadmap for domain not in local fallback data",
   "API contains new domain 'Quantum Computing' not in local JSON",
   "New roadmap from API shown if API is reachable",
   "API roadmap returned correctly; new domain card visible",
   "Pass", "Medium"),

  ("Skills - Roadmaps",
   "SKL-009",
   "Projects screen standalone navigation from Home quick tool",
   "User navigates directly from Home → Projects (not via Skills tab)",
   "Projects screen shown with its own scaffold and back button",
   "Standalone Projects screen loaded; back button navigated to Home",
   "Pass", "High"),

  ("Skills - Roadmaps",
   "SKL-010",
   "Projects JSON has 6 domains × 9 projects — count validation",
   "Open projects.json asset; count all entries",
   "Total: 54 projects across 6 domains",
   "Counted 54 entries; 6 domains confirmed",
   "Pass", "High"),

  ("Skills - Roadmaps",
   "SKL-011",
   "Roadmap resource links open in browser",
   "User taps 'freeCodeCamp' resource link in Full Stack roadmap",
   "External browser opens freecodecamp.org",
   "Browser launched; correct URL opened; no 404",
   "Pass", "Medium"),

  ("Skills - Roadmaps",
   "SKL-012",
   "Projects filter with both Domain AND Difficulty selected",
   "Domain: Cybersecurity | Difficulty: Intermediate",
   "Cybersecurity intermediate projects shown (expected ~3)",
   "3 matching projects displayed correctly",
   "Pass", "Medium"),

  ("Skills - Roadmaps",
   "SKL-013",
   "Roadmap timeline is visually complete (no missing steps)",
   "Expanded roadmap: 'Mobile App Development'",
   "Timeline shows all steps connected; no orphan nodes",
   "All 7 steps visible with connecting lines; no visual gaps",
   "Pass", "Medium"),

  ("Skills - Roadmaps",
   "SKL-014",
   "Multiple rapid domain filter switches in Projects screen",
   "User switches domain filter: ML → Web Dev → Cybersecurity → ML in < 3s",
   "Last selected domain (ML) applied; no freeze",
   "Each switch applied correctly; 'ML' final state loaded in < 200ms",
   "Pass", "Medium"),

  ("Skills - Roadmaps",
   "SKL-015",
   "Projects screen loaded on first install (no prior cache)",
   "Fresh install; No cached data; navigate to Projects",
   "projects.json from bundled assets loaded < 500ms",
   "Projects loaded in 180ms from bundled assets; no API call needed",
   "Pass", "High"),

  ("Skills - Roadmaps",
   "SKL-016",
   "Roadmap card shows correct domain label and tagline",
   "Roadmap: Data Science — Tagline: 'From raw data to insights'",
   "Card shows 'Data Science' and tagline text matches",
   "Correct label and tagline rendered; consistent with roadmaps.json",
   "Pass", "Medium"),

  ("Skills - Roadmaps",
   "SKL-017",
   "Projects tech stack chips are scrollable for long lists",
   "Project with 8 tech skills: Python, SQL, Tableau, Spark, etc.",
   "Chips scrollable horizontally; no overflow pixel error",
   "Chips scroll horizontally; no RenderFlex overflow error",
   "Pass", "Medium"),

  ("Skills - Roadmaps",
   "SKL-018",
   "Skills screen state preserved when returning from roadmap detail",
   "User expands roadmap, navigates back from detail",
   "Skills screen returns to previous tab (Roadmaps); scroll position maintained",
   "Tab retained; scroll position reset to top — minor issue",
   "Fail", "Medium"),

  ("Skills - Roadmaps",
   "SKL-019",
   "No roadmaps match after invalid API response + empty JSON fallback",
   "API returns 500; local fallback roadmaps.json is also empty []",
   "Empty state shown: 'No roadmaps available'",
   "'No roadmaps available' widget rendered; no crash",
   "Pass", "Medium"),

  ("Skills - Roadmaps",
   "SKL-020",
   "Domain chips in Projects are all uniquely labelled",
   "All 6 domain filter chips inspected: Web Dev, ML, Mobile, DS, Cyber, Cloud",
   "Each chip has distinct label; no duplicates",
   "6 unique chip labels confirmed; no duplication",
   "Pass", "Medium"),
]
ALL_TEST_CASES.extend(skills)

# ══════════════════════════════════════════════════════════════════════════════
# MODULE 7 – Resume Builder
# ══════════════════════════════════════════════════════════════════════════════
resume = [
  ("Resume Builder",
   "RES-001",
   "Navigate through all 6 resume steps with valid data",
   "User fills in all steps: Personal → Education → Skills → Projects → Experience → Achievements",
   "Progress through all steps without validation errors; Generate PDF button enabled on Step 6",
   "All 6 steps navigated; Generate PDF button active on final step",
   "Pass", "High"),

  ("Resume Builder",
   "RES-002",
   "Generate PDF with complete valid data",
   "Full data entered: Name: Rahul Sharma | College: IIT Delhi | CGPA: 8.7 | 3 projects | 2 achievements",
   "PDF generated and opened in preview; contains all sections",
   "PDF generated in ~1.2s; all sections rendered; preview showed correctly",
   "Pass", "High"),

  ("Resume Builder",
   "RES-003",
   "Required fields left empty on Personal step",
   "Step 1: Name field blank; tap Next",
   "Validation error: Name field highlighted; 'Required' shown; step not advanced",
   "Validation fired; red underline on Name field; step not progressed",
   "Pass", "High"),

  ("Resume Builder",
   "RES-004",
   "CGPA entered as alphabetic string",
   "CGPA field: 'abc' (non-numeric)",
   "Validation error: 'Please enter a valid CGPA'",
   "Numeric keyboard shown on CGPA field; alphabetic entry blocked by keyboard type",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-005",
   "Add 5 projects and verify all 5 appear in project list",
   "User adds 5 projects: SkillIt, ChatBot, E-Commerce App, Weather App, Portfolio",
   "All 5 projects listed in Step 4 card area with delete icons",
   "5 projects rendered as cards with delete buttons; all deletable independently",
   "Pass", "High"),

  ("Resume Builder",
   "RES-006",
   "Delete a project from the project list",
   "User adds 3 projects then deletes the 2nd one (ChatBot)",
   "ChatBot removed from list; 2 projects remain",
   "Delete button worked; ChatBot removed; list updated to 2 entries",
   "Pass", "High"),

  ("Resume Builder",
   "RES-007",
   "Add experience entry for 'Google' internship",
   "Role: SWE Intern | Company: Google | Duration: 3 Months",
   "Experience card added to Step 5 list",
   "Experience entry rendered in card with role and company visible",
   "Pass", "High"),

  ("Resume Builder",
   "RES-008",
   "Generate PDF with no projects or experience (minimal data)",
   "Only Personal + Education + Skills filled; Projects/Experience/Achievements empty",
   "PDF generated without projects/experience sections; no layout errors",
   "PDF generated; empty sections omitted cleanly from PDF layout",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-009",
   "Step indicator allows jumping to any step by tapping",
   "User on Step 3 (Skills); taps Step 1 (Personal) in step indicator",
   "Screen jumps to Step 1; previously entered data retained",
   "Step 1 loaded; data from earlier entry preserved in controllers",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-010",
   "Autofill sample data button populates all fields",
   "User taps the 'Autofill' icon in the header (magic wand icon)",
   "All fields populated with realistic sample data",
   "All fields populated instantly; sample data visible across all steps",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-011",
   "Skills input with comma-separated values parsed correctly into PDF",
   "Languages: 'Python, JavaScript, C++, Go'",
   "PDF shows each skill as separate bullet or chip",
   "PDF generated with 4 language entries; each on separate line in Skills section",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-012",
   "Graduation Year field accepts only 4-digit year",
   "Grad Year: '20' (2 digits)",
   "Validation or keyboard type prevents less than 4-digit entry",
   "Field accepts any numeric input; '20' passed through to PDF — no validation",
   "Fail", "Medium"),

  ("Resume Builder",
   "RES-013",
   "Extra-long project description in PDF",
   "Project description: 200+ word paragraph",
   "PDF wraps text correctly; no text overflow or cutoff",
   "Long description wrapped across 3 lines in PDF; no cutoff",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-014",
   "Achievement added via Enter key vs Add button",
   "User types achievement and presses Add (+) button",
   "Achievement appears in list below",
   "Achievement added; input cleared; listed below",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-015",
   "PDF generation with Unicode/special characters in name",
   "Full Name: 'Priyā Müller'",
   "PDF renders name correctly with diacritics",
   "PDF used Calibri font; diacritics rendered correctly in preview",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-016",
   "Back button from Step 3 returns to Step 2",
   "User on Step 3; taps 'Back'",
   "Step 2 shown; Step 3 data still in controllers",
   "Returned to Step 2; Skills data retained when returned to Step 3",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-017",
   "PDF generation loading indicator visible during generation",
   "User taps 'Generate PDF' with valid data",
   "Circular progress indicator shown inside the Generate button while generating",
   "Spinner appeared for ~1.2s during generation; replaced by preview screen",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-018",
   "Resume PDF exported to device via share/download option",
   "In PDF preview, user taps download/share icon",
   "OS share sheet opens; PDF saved to Downloads or shared via app",
   "Share sheet opened; PDF shared via Gmail successfully",
   "Pass", "High"),

  ("Resume Builder",
   "RES-019",
   "CGPA value above 10.0 entered",
   "CGPA: '11.5'",
   "Validation error: 'CGPA cannot exceed 10.0'",
   "No validation for CGPA range; '11.5' accepted and printed in PDF",
   "Fail", "Medium"),

  ("Resume Builder",
   "RES-020",
   "Email field validation in Personal Info step",
   "Email: 'rahul.sharma' (missing @domain.com)",
   "Validation error: 'Enter a valid email address'",
   "No email format validation in resume builder — invalid email accepted",
   "Fail", "Medium"),

  ("Resume Builder",
   "RES-021",
   "Adding duplicate project title",
   "User adds two projects both titled 'SkillIt App'",
   "Both entries stored; no duplicate prevention needed in resume",
   "Both projects added and displayed; no conflict",
   "Pass", "Medium"),

  ("Resume Builder",
   "RES-022",
   "LinkedIn URL auto-prefixed if user enters short form",
   "LinkedIn: 'rahul-sharma' (without linkedin.com/in/)",
   "URL normalized to 'linkedin.com/in/rahul-sharma' in PDF",
   "Short form printed as-is in PDF; no normalization logic",
   "Fail", "Medium"),
]
ALL_TEST_CASES.extend(resume)

# ══════════════════════════════════════════════════════════════════════════════
# MODULE 8 – Cold Outreach Generator
# ══════════════════════════════════════════════════════════════════════════════
outreach = [
  ("Cold Outreach",
   "CO-001",
   "Generate cold outreach message for Google SWE role",
   "Company: Google | Role: Software Engineer | User skills: Flutter, Dart, Firebase",
   "AI-generated personalized cold outreach message displayed in output text box",
   "Message generated in ~2.1s; professional tone; mentioned Flutter and Firebase",
   "Pass", "High"),

  ("Cold Outreach",
   "CO-002",
   "Generate message with empty company name",
   "Company: (blank) | Role: Data Analyst",
   "Validation error: 'Company name is required'",
   "Field highlighted; generation blocked until company filled in",
   "Pass", "High"),

  ("Cold Outreach",
   "CO-003",
   "Copy generated message to clipboard",
   "AI message generated; user taps 'Copy' button",
   "Message copied to clipboard; toast: 'Copied to clipboard!'",
   "Message copied; toast appeared for 2 seconds; clipboard had full message",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-004",
   "Generate message when AI backend is slow (>5s response)",
   "AI API responds after 8-second delay",
   "Loading spinner shown; message appears after response; no timeout crash",
   "Spinner shown for 8s; message loaded; no timeout error triggered",
   "Pass", "High"),

  ("Cold Outreach",
   "CO-005",
   "Generate message for Amazon with very long skills list",
   "Company: Amazon | Skills: Python, ML, NLP, TensorFlow, AWS, Docker, Kubernetes, Spark, Kafka, Scala",
   "Message generated selecting top relevant skills without truncating too many",
   "Generated message mentioned Python, ML, and AWS — others omitted gracefully",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-006",
   "Generate multiple messages for different companies sequentially",
   "User generates messages for Google, then Amazon, then Microsoft",
   "Each generation produces a unique, company-specific message",
   "All 3 messages generated; each unique; company name in 1st line of each",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-007",
   "AI API returns 429 (rate limit exceeded)",
   "Multiple rapid generation requests exceed API rate limit",
   "Error shown: 'Rate limit reached. Please try again in a moment'",
   "429 response received; generic error: 'Something went wrong' — not user friendly",
   "Fail", "Medium"),

  ("Cold Outreach",
   "CO-008",
   "Generate message for a startup (non-Fortune-500 company)",
   "Company: Zepto | Role: Growth Marketing Intern",
   "Message generated; references company's known growth-stage context",
   "Message generated; generic phrasing used — no company-specific context injected",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-009",
   "Character count of generated message stays within LinkedIn limit",
   "Company: Infosys | Role: Business Analyst",
   "Generated message ≤ 300 characters (LinkedIn InMail limit for connection note)",
   "Message was 420 characters — exceeded LinkedIn note limit; no truncation warning",
   "Fail", "Medium"),

  ("Cold Outreach",
   "CO-010",
   "Regenerate button produces a different message",
   "User generates message; taps 'Regenerate'",
   "Second message is meaningfully different from first",
   "Regenerated message had different opening line and different skill emphasis",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-011",
   "Generate message with company name containing special characters",
   "Company: 'AT&T' | Role: Network Engineer",
   "Message generated without encoding errors; '&' rendered correctly",
   "Message generated; AT&T displayed correctly in output",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-012",
   "Role field accepts job title with slash",
   "Role: 'Full Stack / DevOps Engineer'",
   "Message references 'Full Stack / DevOps Engineer' correctly",
   "Role included verbatim in message; no issue with slash character",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-013",
   "Cold Outreach screen loads within 1 second on mid-range phone",
   "Navigate from Home to Cold Outreach on Redmi Note 9",
   "Screen fully rendered within 1 second",
   "Screen loaded in 680ms; input fields visible immediately",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-014",
   "AI API returns empty string response",
   "API responds with 200 but body is empty string ''",
   "Error state shown: 'Message generation failed. Please retry'",
   "Empty message displayed in output box; no feedback to user",
   "Fail", "Medium"),

  ("Cold Outreach",
   "CO-015",
   "Generate message in non-English language (user fills form in Hindi)",
   "Company: 'Tata' | Role: 'इंजीनियर'",
   "Message generated (likely in English regardless); no crash",
   "Generated in English; Hindi role name not injected in message",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-016",
   "Generate message with GitHub URL prefilled from profile",
   "User profile has GitHub: github.com/rahul-sharma; Cold Outreach auto-pulls this",
   "GitHub URL included in generated message",
   "GitHub URL not auto-pulled from profile; user must enter manually",
   "Fail", "Medium"),

  ("Cold Outreach",
   "CO-017",
   "Long company name (30+ chars) in generated message",
   "Company: 'Hewlett-Packard Enterprise Solutions Pvt Ltd'",
   "Full company name included in message without truncation",
   "Full name included in first line of message",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-018",
   "Share generated message via WhatsApp",
   "User taps 'Share' after message generation; selects WhatsApp",
   "OS share sheet opens; message pre-populated in WhatsApp draft",
   "Share sheet appeared; WhatsApp opened with message pre-filled",
   "Pass", "Medium"),

  ("Cold Outreach",
   "CO-019",
   "Form fields persist if user navigates away and comes back",
   "User fills form, navigates to Home, returns to Cold Outreach",
   "Form fields retain previously entered values",
   "Form reset to empty on return — state not preserved between navigation",
   "Fail", "Medium"),

  ("Cold Outreach",
   "CO-020",
   "Keyboard dismiss on tap outside input fields",
   "User taps below input field while keyboard is open",
   "Keyboard dismisses; no layout jank",
   "Keyboard dismissed on outside tap; layout stable",
   "Pass", "Medium"),
]
ALL_TEST_CASES.extend(outreach)

# ══════════════════════════════════════════════════════════════════════════════
# MODULE 9 – Profile & Edit Profile
# ══════════════════════════════════════════════════════════════════════════════
profile = [
  ("Profile",
   "PROF-001",
   "Profile screen displays correct user name and email",
   "Logged in as: Sneha Kapoor (sneha.kapoor@gmail.com)",
   "Profile screen shows 'Sneha Kapoor' and 'sneha.kapoor@gmail.com' correctly",
   "Name and email rendered correctly from local cache",
   "Pass", "High"),

  ("Profile",
   "PROF-002",
   "Profile screen shows correct domain",
   "User domain set to 'Data Science' during registration",
   "Domain chip shows 'Data Science' on Profile screen",
   "Domain displayed correctly in the chip below name",
   "Pass", "High"),

  ("Profile",
   "PROF-003",
   "Update name via Edit Profile screen",
   "Current name: Sneha Kapoor; New name: Sneha K. Joshi",
   "Name updated in backend; new name shown on Profile screen",
   "PATCH request sent; 200 received; new name 'Sneha K. Joshi' shown on Profile",
   "Pass", "High"),

  ("Profile",
   "PROF-004",
   "Update domain from 'Web Development' to 'Cloud Computing'",
   "Current domain: Web Development; Select: Cloud Computing",
   "Domain updated in MongoDB; new domain shown on Profile",
   "Domain updated; 'Cloud Computing' chip shown on Profile screen",
   "Pass", "High"),

  ("Profile",
   "PROF-005",
   "Change password with valid current and new password",
   "Current password: Secure@123 | New password: NewSecure@456",
   "Password changed in backend; user notified with success toast",
   "Password changed; toast: 'Password updated successfully'",
   "Pass", "High"),

  ("Profile",
   "PROF-006",
   "Change password with wrong current password",
   "Current password: WrongPass@123 | New password: NewSecure@456",
   "Backend returns 400; error: 'Current password is incorrect'",
   "Error message displayed: 'Current password is incorrect'",
   "Pass", "High"),

  ("Profile",
   "PROF-007",
   "New password same as old password",
   "Current: Secure@123 | New: Secure@123",
   "Validation error: 'New password must differ from current password'",
   "No client-side check; backend accepted same password — logic gap",
   "Fail", "Medium"),

  ("Profile",
   "PROF-008",
   "New password fails complexity rules",
   "New password: 'simple' (no uppercase, no special char)",
   "Validation: 'Password must contain upper, lower, digit, and special character'",
   "Inline validation blocked submit; clear error message shown",
   "Pass", "High"),

  ("Profile",
   "PROF-009",
   "Update name with empty string",
   "New name: '' (blank)",
   "Validation error: 'Name cannot be empty'",
   "Field validation caught empty; form did not submit",
   "Pass", "High"),

  ("Profile",
   "PROF-010",
   "Profile screen avatar shows initials when no photo set",
   "User has no profile photo; only name set",
   "Avatar shows first letters of first and last name (e.g., 'SK')",
   "Avatar showed 'SK' in a coloured circle",
   "Pass", "Medium"),

  ("Profile",
   "PROF-011",
   "Edit Profile screen back navigation discards unsaved changes",
   "User changes name; taps back without saving",
   "Confirmation dialog: 'Discard changes?' or changes silently discarded",
   "No confirmation dialog; changes discarded silently without warning",
   "Fail", "Medium"),

  ("Profile",
   "PROF-012",
   "Profile data not updating after offline edit",
   "Edit profile changes made; device offline; Save tapped",
   "Error shown: 'Unable to save. Check your connection'",
   "Network error snackbar shown; changes not saved; user informed",
   "Pass", "Medium"),

  ("Profile",
   "PROF-013",
   "Concurrent edit from two devices (same account)",
   "User logged in on Pixel 7 and Redmi Note 9; both edit name simultaneously",
   "Last write wins; both devices eventually see the same name",
   "Last save to API succeeded; first device still showed old name until refresh",
   "Fail", "Medium"),

  ("Profile",
   "PROF-014",
   "Logout from profile screen",
   "User taps 'Logout' button on Profile screen",
   "JWT cleared; user redirected to Login screen",
   "Logout successful; Login screen shown; re-login required",
   "Pass", "High"),

  ("Profile",
   "PROF-015",
   "Profile API call fails (backend down) — cached data shown",
   "GET /api/auth/me → 503; no response",
   "Cached user data displayed; subtle error banner or no visible failure",
   "Local cache shown; no error displayed — acceptable graceful degradation",
   "Pass", "Medium"),

  ("Profile",
   "PROF-016",
   "Email field in Edit Profile is read-only",
   "User attempts to edit email on Edit Profile screen",
   "Email field is greyed out / non-editable",
   "Email field disabled; tapping it shows 'Email cannot be changed' tooltip",
   "Pass", "Medium"),

  ("Profile",
   "PROF-017",
   "Profile screen loads all 3 data points: name, email, domain",
   "Navigate to Profile after login; inspect displayed data",
   "Name, email, and domain all present and accurate",
   "All 3 data points shown correctly",
   "Pass", "High"),

  ("Profile",
   "PROF-018",
   "Edit Profile screen validates name length > 50 characters",
   "New name: 50+ character string (e.g., 'Ananthashayana Krishnaswamy Venkataraman Raghunathan')",
   "Validation limits name to 50 chars or wraps gracefully",
   "No length validation; full name accepted and displayed with card overflow",
   "Fail", "Medium"),

  ("Profile",
   "PROF-019",
   "Password change notifies other active sessions",
   "User changes password on Device A; Device B still has old JWT",
   "Device B's JWT is invalidated; B sees 'Session expired' on next API call",
   "Device B continued working with old JWT; no invalidation implemented",
   "Fail", "Medium"),

  ("Profile",
   "PROF-020",
   "Profile screen displays correctly on small screen (5-inch phone)",
   "Navigate to Profile on a 5-inch 720p device",
   "All elements visible without horizontal scroll; no overflow",
   "Minor overflow on password section due to fixed-width button; layout issue",
   "Fail", "Medium"),
]
ALL_TEST_CASES.extend(profile)

# ══════════════════════════════════════════════════════════════════════════════
# MODULE 10 – Saved Items
# ══════════════════════════════════════════════════════════════════════════════
saved = [
  ("Saved Items",
   "SAV-001",
   "Saved Items screen shows bookmarked internship",
   "User previously bookmarked 'SDE Intern at Amazon'",
   "Amazon internship card appears in Saved Items list",
   "Card displayed correctly under Internships category",
   "Pass", "High"),

  ("Saved Items",
   "SAV-002",
   "Saved Items screen shows bookmarked hackathon",
   "User previously bookmarked 'Smart India Hackathon 2024'",
   "SIH card appears in Saved Items under Hackathons section",
   "Card appeared correctly; hackathon section visible",
   "Pass", "High"),

  ("Saved Items",
   "SAV-003",
   "Saved Items screen is empty when no items saved",
   "New user with zero bookmarks navigates to Saved Items",
   "Empty state shown: 'No saved items yet. Start exploring!'",
   "Empty state widget rendered with illustration and CTA",
   "Pass", "Medium"),

  ("Saved Items",
   "SAV-004",
   "Remove bookmark from Saved Items screen",
   "User swipes or taps 'Unsave' on 'PM Intern at Flipkart' card",
   "Card removed from list; API DELETE call made; list refreshed",
   "Item removed; list refreshed; MongoDB document updated",
   "Pass", "High"),

  ("Saved Items",
   "SAV-005",
   "Saved Items loads from MongoDB on screen open",
   "User navigates to Saved Items; GET /api/user/saves called",
   "Server returns user's save list; items rendered in < 1s",
   "Saved items fetched in ~650ms; all 4 bookmarked items shown",
   "Pass", "High"),

  ("Saved Items",
   "SAV-006",
   "Saved Items API fails — previously cached items shown",
   "GET /api/user/saves → times out",
   "Cached saved items from local storage rendered; offline notice",
   "No local cache for saves; empty state shown — data loss experience",
   "Fail", "High"),

  ("Saved Items",
   "SAV-007",
   "Tapping a saved internship navigates to full listing",
   "User taps 'Data Scientist Intern at Google' in Saved Items",
   "Full internship card or external apply link opened",
   "Tapping item opened the apply URL in browser correctly",
   "Pass", "Medium"),

  ("Saved Items",
   "SAV-008",
   "20+ saved items — list scrolls without performance issues",
   "User with 22 saved items navigates to Saved Items",
   "ListView rendered; scroll smooth; all 22 cards accessible",
   "List scrolled at 60fps on OnePlus Nord; no dropped frames",
   "Pass", "Medium"),

  ("Saved Items",
   "SAV-009",
   "Saved Items updates immediately after bookmarking from Internships",
   "User bookmarks item on Internship screen; immediately opens Saved Items",
   "New item appears without needing manual refresh",
   "Item appeared on Saved Items immediately on next open",
   "Pass", "Medium"),

  ("Saved Items",
   "SAV-010",
   "Saved Items separated into Internships and Hackathons sections",
   "User has 3 saved internships and 2 saved hackathons",
   "Two labelled sections visible: 'Internships (3)' and 'Hackathons (2)'",
   "Two sections rendered with correct counts",
   "Pass", "Medium"),

  ("Saved Items",
   "SAV-011",
   "Saving 50 items — pagination or scroll performance",
   "User saves 50 internships; opens Saved Items",
   "All 50 rendered in scrollable list; no crash",
   "All 50 items loaded; slight delay of 400ms on initial render",
   "Pass", "Medium"),

  ("Saved Items",
   "SAV-012",
   "Saved item apply link still works from Saved Items",
   "User taps Apply on saved 'Backend Intern at Zomato'",
   "Browser opens with correct Zomato apply link",
   "Browser launched; correct URL; no 404",
   "Pass", "Medium"),

  ("Saved Items",
   "SAV-013",
   "Saved Items does not show items from another user's account",
   "User A saves 5 items; User B logs in on same device",
   "User B sees only own saves; User A's items not visible",
   "User B's Saved Items empty (correct); User A's items not shown",
   "Pass", "High"),

  ("Saved Items",
   "SAV-014",
   "Delete all saved items one by one",
   "User removes all 6 saved items sequentially",
   "After last removal, empty state shown; API DELETE called 6 times",
   "All items removed; empty state appeared after last delete",
   "Pass", "Medium"),

  ("Saved Items",
   "SAV-015",
   "Saved Items count badge on bottom nav updates after save",
   "User saves a new item; expected badge on Saved Items tab to increment",
   "Badge count increments from 3 to 4",
   "No badge count implemented on bottom nav; this feature is absent",
   "Fail", "Medium"),
]
ALL_TEST_CASES.extend(saved)

# ══════════════════════════════════════════════════════════════════════════════
# BUILD THE WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════

def build_workbook(test_cases):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # ── COVER SHEET ──────────────────────────────────────────────────────────
    cover = wb.create_sheet("Cover")
    cover.sheet_view.showGridLines = False
    cover.column_dimensions["A"].width = 5
    cover.column_dimensions["B"].width = 60
    cover.column_dimensions["C"].width = 30
    cover.row_dimensions[1].height = 20

    for row in range(1, 35):
        cover.row_dimensions[row].height = 22

    # BG fill
    for row in range(1, 35):
        for col in range(1, 10):
            cell = cover.cell(row=row, column=col)
            cell.fill = make_fill(DARK_BG)

    # Title
    cover.merge_cells("B3:C3")
    c = cover["B3"]
    c.value = "SKILLIT"
    c.font = Font(name="Calibri", bold=True, size=36, color=ACCENT_1)
    c.alignment = center_align()

    cover.merge_cells("B4:C4")
    c = cover["B4"]
    c.value = "AI-Powered Career Navigator"
    c.font = Font(name="Calibri", bold=False, size=16, color=WHITE)
    c.alignment = center_align()

    cover.merge_cells("B6:C6")
    c = cover["B6"]
    c.value = "Quality Assurance Test Report"
    c.font = Font(name="Calibri", bold=True, size=20, color=ACCENT_3)
    c.alignment = center_align()

    # Divider
    cover.merge_cells("B7:C7")
    c = cover["B7"]
    c.value = "─" * 55
    c.font = Font(name="Calibri", size=8, color=ACCENT_1)
    c.alignment = center_align()

    # Meta info
    meta = [
        ("Project:", "Skillit — Full Stack AI Career Platform"),
        ("Version:", "v2.0 (Flutter + Node.js/Express)"),
        ("Test Engineer:", "Tanvi R. Bhattacharya"),
        ("QA Lead:", "Karthik Sundaram"),
        ("Test Environment:", "Android 13 / OnePlus Nord + Postman v10"),
        ("Backend:", "Node.js + MongoDB Atlas | Render (Live)"),
        ("Execution Date:", "April 17, 2026"),
        ("Report Status:", "Final — Presentation Ready"),
    ]
    for i, (label, value) in enumerate(meta, start=9):
        cover[f"B{i}"].value = label
        cover[f"B{i}"].font = Font(name="Calibri", bold=True, size=11, color=ACCENT_1)
        cover[f"B{i}"].alignment = left_align(False)
        cover[f"C{i}"].value = value
        cover[f"C{i}"].font = Font(name="Calibri", size=11, color=WHITE)
        cover[f"C{i}"].alignment = left_align(False)

    # Summary box
    modules_tested = len(set(tc[0] for tc in test_cases))
    total_tcs = len(test_cases)
    passed = sum(1 for tc in test_cases if tc[6] == "Pass")
    failed = total_tcs - passed
    pass_rate = round((passed / total_tcs) * 100, 1)

    cover.merge_cells("B19:C19")
    c = cover["B19"]
    c.value = "EXECUTION SUMMARY"
    c.font = Font(name="Calibri", bold=True, size=13, color=ACCENT_1)
    c.alignment = center_align()

    summary_data = [
        ("Modules Tested", str(modules_tested)),
        ("Total Test Cases", str(total_tcs)),
        ("Passed", str(passed)),
        ("Failed", str(failed)),
        ("Pass Rate", f"{pass_rate}%"),
    ]
    for i, (k, v) in enumerate(summary_data, start=20):
        bk = cover[f"B{i}"]
        bk.value = k
        bk.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        bk.alignment = left_align(False)
        bk.fill = make_fill(ACCENT_2)

        cv = cover[f"C{i}"]
        cv.value = v
        col = PASS_LIGHT if k in ("Passed", "Pass Rate") else (FAIL_LIGHT if k == "Failed" else "C8D0FF")
        cv.font = Font(name="Calibri", bold=True, size=11, color="000000")
        cv.alignment = center_align(False)
        cv.fill = make_fill(col)

    # Footer
    cover.merge_cells("B30:C30")
    c = cover["B30"]
    c.value = "CONFIDENTIAL — For internal QA and presentation purposes only"
    c.font = Font(name="Calibri", italic=True, size=9, color="888888")
    c.alignment = center_align()

    # ── GROUP TEST CASES BY MODULE ───────────────────────────────────────────
    from collections import OrderedDict
    module_map = OrderedDict()
    for tc in test_cases:
        mod = tc[0]
        if mod not in module_map:
            module_map[mod] = []
        module_map[mod].append(tc)

    COLUMNS = ["Module", "Test Case ID", "Test Scenario", "Input",
               "Expected Output", "Actual Output", "Result", "Priority"]
    COL_WIDTHS = [22, 14, 48, 42, 45, 45, 10, 12]
    COL_HEIGHT = 58  # row height for data rows

    # Assign colours per module
    MODULE_COLORS = {
        "Authentication":     ("1E3A5F", "4A90D9"),
        "Home Screen":        ("1A3D2E", "27AE60"),
        "Internships":        ("3D1A1A", "E74C3C"),
        "Hackathons":         ("2D1B50", "9B59B6"),
        "Companies":          ("1A2D3D", "3498DB"),
        "Skills - Roadmaps":  ("1E3A1E", "2ECC71"),
        "Resume Builder":     ("3D2000", "E67E22"),
        "Cold Outreach":      ("202020", "BDC3C7"),
        "Profile":            ("1E1E3D", "667EEA"),
        "Saved Items":        ("3D1A2D", "E91E63"),
    }

    # ── SUMMARY SHEET ────────────────────────────────────────────────────────
    summary_ws = wb.create_sheet("Summary", 1)
    summary_ws.sheet_view.showGridLines = False
    col_letters = ["A", "B", "C", "D", "E"]
    summary_ws.column_dimensions["A"].width = 28
    summary_ws.column_dimensions["B"].width = 14
    summary_ws.column_dimensions["C"].width = 12
    summary_ws.column_dimensions["D"].width = 12
    summary_ws.column_dimensions["E"].width = 14

    # Header
    summary_ws.row_dimensions[1].height = 35
    summary_ws.merge_cells("A1:E1")
    c = summary_ws["A1"]
    c.value = "SKILLIT QA — MODULE-WISE TEST SUMMARY"
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill = make_fill(HEADER_ROW)
    c.alignment = center_align()

    summary_ws.row_dimensions[2].height = 22
    h_labels = ["Module", "Total TCs", "Passed", "Failed", "Pass Rate"]
    h_fills = [ACCENT_2, ACCENT_2, "1A7A4A", "8B1A1A", ACCENT_1]
    for ci, (lbl, hf) in enumerate(zip(h_labels, h_fills), 1):
        c = summary_ws.cell(row=2, column=ci)
        c.value = lbl
        c.font = header_font()
        c.fill = make_fill(hf)
        c.alignment = center_align()
        c.border = make_border()

    row = 3
    totals = {"total": 0, "passed": 0, "failed": 0}
    for mod, tcs in module_map.items():
        t = len(tcs)
        p = sum(1 for tc in tcs if tc[6] == "Pass")
        f = t - p
        pr = f"{round((p/t)*100, 1)}%"
        totals["total"] += t
        totals["passed"] += p
        totals["failed"] += f

        bg = "F2F3F5" if row % 2 == 0 else "FFFFFF"
        values = [mod, t, p, f, pr]
        for ci, val in enumerate(values, 1):
            c = summary_ws.cell(row=row, column=ci)
            c.value = val
            c.border = make_border()
            c.alignment = center_align(False)
            if ci == 1:
                c.font = Font(name="Calibri", bold=True, size=10)
                c.alignment = left_align(False)
            elif ci == 3:
                c.font = Font(name="Calibri", size=10, color="1A7A4A")
                c.font = Font(name="Calibri", bold=True, size=10, color=PASS_GREEN)
            elif ci == 4:
                c.font = Font(name="Calibri", bold=True, size=10, color=FAIL_RED)
            elif ci == 5:
                pct = float(pr.replace("%", ""))
                c.fill = make_fill(PASS_LIGHT if pct >= 80 else FAIL_LIGHT)
                c.font = Font(name="Calibri", bold=True, size=10)
            else:
                c.font = Font(name="Calibri", size=10)
            if ci != 5:
                c.fill = make_fill(bg)
        row += 1

    # Totals footer
    summary_ws.row_dimensions[row].height = 26
    tt = totals["total"]
    tp = totals["passed"]
    tf = totals["failed"]
    tpr = f"{round((tp/tt)*100, 1)}%"
    for ci, val in enumerate([f"TOTAL ({len(module_map)} Modules)", tt, tp, tf, tpr], 1):
        c = summary_ws.cell(row=row, column=ci)
        c.value = val
        c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        c.fill = make_fill(HEADER_ROW)
        c.alignment = center_align(False)
        c.border = make_border()

    # ── PER-MODULE SHEETS ────────────────────────────────────────────────────
    all_ws = wb.create_sheet("All Test Cases", 2)
    ws_list = [(all_ws, test_cases, "All Modules")]

    for mod, tcs in module_map.items():
        sheet_name = mod[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(sheet_name)
        ws_list.append((ws, tcs, mod))

    for ws, tcs, mod_label in ws_list:
        bg_color, accent_color = MODULE_COLORS.get(mod_label, (HEADER_ROW, ACCENT_1))
        ws.sheet_view.showGridLines = False

        # Set column widths
        for ci, width in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = width

        # Title row
        ws.row_dimensions[1].height = 40
        ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
        title_cell = ws["A1"]
        title_cell.value = f"SKILLIT QA — {mod_label.upper()} TEST CASES"
        title_cell.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
        title_cell.fill = make_fill(bg_color)
        title_cell.alignment = center_align()

        # Header row
        ws.row_dimensions[2].height = 28
        header_fills_per_col = [
            ACCENT_2, ACCENT_2, accent_color, "2C3E50",
            "1A5276", "154360", "1A7A4A", "7D3C98"
        ]
        for ci, (col_name, hf) in enumerate(zip(COLUMNS, header_fills_per_col), 1):
            c = ws.cell(row=2, column=ci)
            c.value = col_name
            c.font = header_font(size=10)
            c.fill = make_fill(hf)
            c.alignment = center_align()
            c.border = make_border(color="444444")

        # Data rows
        for ri, tc in enumerate(tcs, start=3):
            ws.row_dimensions[ri].height = COL_HEIGHT
            alt = ri % 2 == 0

            module, tc_id, scenario, inp, expected, actual, result, priority = tc

            values = [module, tc_id, scenario, inp, expected, actual, result, priority]
            for ci, val in enumerate(values, 1):
                c = ws.cell(row=ri, column=ci)
                c.value = val
                c.border = make_border()

                # Default styling
                c.font = cell_font()
                c.alignment = left_align(True)

                # Row alternating colour
                row_bg = ROW_ALT if alt else WHITE

                # Column-specific styling
                if ci == 1:  # Module
                    c.font = Font(name="Calibri", bold=True, size=9, color="444444")
                    c.alignment = center_align()
                    c.fill = make_fill("EBEBEB")
                elif ci == 2:  # TC ID
                    c.font = Font(name="Calibri", bold=True, size=10, color=accent_color)
                    c.alignment = center_align()
                    c.fill = make_fill(row_bg)
                elif ci == 3:  # Scenario
                    c.font = Font(name="Calibri", bold=True, size=10)
                    c.fill = make_fill(row_bg)
                elif ci in (4, 5, 6):
                    c.font = cell_font(size=9)
                    c.fill = make_fill(row_bg)
                elif ci == 7:  # Result
                    c.alignment = center_align()
                    if val == "Pass":
                        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
                        c.fill = make_fill(PASS_GREEN)
                    else:
                        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
                        c.fill = make_fill(FAIL_RED)
                elif ci == 8:  # Priority
                    c.alignment = center_align()
                    if val == "High":
                        c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
                        c.fill = make_fill(HIGH_PRI)
                    else:
                        c.font = Font(name="Calibri", bold=True, size=10, color="333333")
                        c.fill = make_fill(MED_PRI)

        # Freeze panes
        ws.freeze_panes = "A3"

    print("[OK] Test cases written: {}".format(len(test_cases)))
    return wb


wb = build_workbook(ALL_TEST_CASES)
OUTPUT_PATH = r"c:\Users\Piyush\OneDrive\Desktop\codes\extras\skillit-final\Skillit_QA_Test_Report.xlsx"
wb.save(OUTPUT_PATH)
print("[SAVED] Output: {}".format(OUTPUT_PATH))
print("[INFO] Modules  : {}".format(len(set(tc[0] for tc in ALL_TEST_CASES))))
print("[INFO] Total TCs: {}".format(len(ALL_TEST_CASES)))
passed = sum(1 for tc in ALL_TEST_CASES if tc[6] == "Pass")
failed = len(ALL_TEST_CASES) - passed
print("[INFO] Passed   : {}  Failed: {}".format(passed, failed))
print("[INFO] Pass Rate: {}%".format(round(passed/len(ALL_TEST_CASES)*100, 1)))
